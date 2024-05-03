
import mailbox
from email.utils import parsedate_to_datetime
import datetime
from datetime import datetime
from email.header import decode_header
import re
from openpyxl import Workbook
import GUI
from GUI import countUp

Mailget_Progress = 0    #進捗表示用

def extract_blocks(text):
    blocks = []
    current_block = []
    in_block = False
    for line in text.splitlines():
        if line.startswith("-----"):
            if in_block:
                blocks.append("".join(current_block))
                current_block = []
            in_block = not in_block
        if in_block:
            current_block.append(line + '\r\n')  # 改行を追加
    return blocks

def contains_digits_and_commas(s):
    contains_digit = False
    contains_comma = False
    for char in s:
        if char.isdigit():
            contains_digit = True
        elif char == ',' or char == '+' or char == '-':
            contains_comma = True
        if contains_digit and contains_comma:
            return True
    return False

def add_backorder_to_cancelled(blocks):
    modified_blocks = []
    for block in blocks:
        if 'キャンセル' in block:
            lines = block.splitlines()
            for i in range(len(lines)):
                search_text = lines[i]

                if contains_digits_and_commas(search_text):
                    lines[i] += ",欠品"
            block = "\r\n".join(lines)
        modified_blocks.append(block)
    return modified_blocks

def remove_lines_with_keywords(text):
    lines = text.split('\n')
    cleaned_lines = [line for line in lines if 'キャンセル' not in line and '本日中' not in line and '。' not in line]
    return '\n'.join(cleaned_lines)

def convert_string(input_string):   #'/'入れ忘れを修正する
    pattern = r'(-|\+)(\d{1,2}\.\d{1,2})(-|\+)(\d{1,2}\.\d{1,2})'
    result = re.sub(pattern, r'\1\2/\3\4', input_string)
    return result

def contains_hiragana_or_katakana(text):
    pattern = re.compile(r'[\u3040-\u309F\u30A0-\u30FF]+')  # ひらがなとカタカナのUnicode範囲
    return bool(re.search(pattern, text))

def remove_non_numeric(input_string):
    pattern = r'[^\d.\+\-]'  # 数字とピリオド,+-以外の文字を表す正規表現
    result = re.sub(pattern, '', input_string)
    return result

def MailGet_func_thunder():
    import main
    if main.selected_directory:
        mail_box = mailbox.mbox(main.selected_directory)
    else:
        raise FileNotFoundError
    end_date = datetime.today().replace(hour=23, minute=59, second=59, microsecond=99)
    start_date = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    to_sort = end_date
    from_sort = start_date

    mail_progress_per = 1000

    try:
        end_date = GUI.to_cal.get()
        start_date = GUI.from_cal.get()
        date_format = "%Y/%m/%d"

        to_sort = datetime.strptime(end_date, date_format).replace(hour=23, minute=59, second=59, microsecond=99)
        from_sort = datetime.strptime(start_date, date_format).replace(hour=0, minute=0, second=0, microsecond=0)

    except Exception as e:
        print('日付指定がありません')
    try:
        indivisual_prog = mail_progress_per / len(mail_box.keys())
    except Exception as e:
        print('受信したメールがありません')

    item_list = ""
    j = 0
    get_mail_amount = 0
    print(len(mail_box.keys()))
    print(mail_box.keys())

    for date, key in zip(mail_box, mail_box.keys()):
        countUp(indivisual_prog)
        j += 1

        email_date = parsedate_to_datetime(date['Date'])
        email_date = email_date.replace(tzinfo=None)
        
        if from_sort <= email_date <= to_sort:

            get_mail_amount += 1

            a_msg = mail_box.get(key)
            from_str = a_msg.get_from()
            print(from_str)

            #usbj = ''
            #for bstr,enc in decode_header(a_msg['Subject']) :
            #    if enc == None:
            #        usbj += bstr #.decode("ascii", "ignore")
            #    else:
            #        usbj += bstr.decode(enc, "ignore")

            for aa_msg in a_msg.walk():
                if not 'text' in aa_msg.get_content_type():
                    continue #"text"パートでなかったら次のパートへ
                if aa_msg.get_content_charset() :
                    print(aa_msg.get_content_charset())
                    a_text = aa_msg.get_payload(decode=True).decode(aa_msg.get_content_charset(), "ignore")
                else:
                    if "charset=shift_jis" in str(aa_msg.get_payload(decode=True)):
                        #ひとまず シフトJISだけ特別対応。
                        a_text = aa_msg.get_payload(decode=True).decode("cp932", "ignore")
                    else:
                        print ("** Cannot decode.Cannot specify charset ***"+aa_msg.get("From"))
                        continue

                if not 'html' in a_text and ('欠品' in a_text or '解除' in a_text):

                    print(a_text)

                    modifyed_text = re.sub(r'\n\s*\n', '\n', a_text)

                    modifyed_text_converted = convert_string(modifyed_text)   #'/'入れ忘れを修正

                    replaced_text = modifyed_text_converted.replace("/",",")#.replace("　",",")

                    matches = re.findall(r'(-----\n.*?\n-----)', replaced_text, re.DOTALL)

                    # 抽出した部分を出力
                    for match in matches:
                        
                        # ブロックを抽出
                        blocks = extract_blocks(match)

                        # レンズ情報に',欠品'を追加
                        modified_blocks = add_backorder_to_cancelled(blocks)

                        # 出力
                        modifyed_blocks_text = "\n-----------------------------------\n".join(modified_blocks)

                        # レンズ情報以外の文言を削除
                        cleaned_text = remove_lines_with_keywords(modifyed_blocks_text)

                        matches2 = re.findall(r'(\n.*?\r)', cleaned_text, re.DOTALL)

                        itemname = matches2[0].replace("\r",",")

                        for k, match2 in enumerate(matches2):
                            if k == 0 : continue
                            output_text = (itemname+ match2).replace("\n","")
                            print(output_text)
                            item_list += output_text

                        print(cleaned_text)

        if j == len(mail_box.keys()) or j == 10000: 
            print(item_list)
            data = item_list.strip().split("\r")

            # Workbookオブジェクトを作成
            wb = Workbook()

            # アクティブなシートを取得
            ws = wb.active

            # 最初の行をヘッダー行として設定
            header = ["商品名", "度数", "乱視度数", "乱視軸","BC","加入度数","DIA","カラー","欠品"]
            for col_num, value in enumerate(header, start=1):
                ws.cell(row=1, column=col_num, value=value)

            # データを書き込む
            for row_num, row_data in enumerate(data, start=2):
                row_values = row_data.split(",")
                for value_num, value_name in enumerate(row_values):
                    if value_num != 0:
                        #商品名以外でスペースの区切りを用いていた場合、スプリットする
                        if " " in value_name or "　" in value_name:
                            modify_value = re.split("[\\s]",value_name)
                            print(modify_value)
                            del row_values[value_num]
                            for modify_num, modify_name in enumerate(modify_value):
                                row_values.insert(value_num + modify_num, modify_name)

                minus = 0

                for col_num, value in enumerate(row_values, start=1):
                    ex_col = False

                    if "欠品" in value:                                         #欠品の表記は9列目
                        ex_col = 9
                    if  col_num != 1 and contains_hiragana_or_katakana(value):  #商品名以外のひらがなカタカナはカラーとみなす、8列目
                        ex_col = 8
                    if "DIA" in value:                                          #DIAは7列目
                        contain = value
                        value = remove_non_numeric(contain)
                        ex_col = 7
                    if "ADD" in value:                                          #ADDは6列目
                        contain = value
                        value = remove_non_numeric(contain)
                        ex_col = 6
                    if "BC" in value:                                           #BCは5列目
                        ex_col = 5
                    if 5 <= ex_col <= 8 and col_num > 1:                        #度数に空白列が出来たら詰める
                        minus = 1

                    if col_num == 2 and ex_col != 8:        #度数の余計な文字列を削除する
                        contain = value
                        value = remove_non_numeric(contain)

                    ws.cell(row=row_num, column= ex_col or col_num - minus, value=value)

            # Excelファイルに保存
            today = format(datetime.today(), '%Y%m%d')
            wb.save(f"mail_import/import_{today}.xlsx")

            if j >= 10000:
                    print('メールの取得件数が10,000件を超えました')

            break
    
    if get_mail_amount == 0:
        raise ZeroDivisionError