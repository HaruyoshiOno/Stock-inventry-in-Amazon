import pandas as pd
from openpyxl import Workbook, load_workbook
import datetime
import GUI
from GUI import countUp
import numpy as np

#本日の日付を取得
today = format(datetime.date.today(), '%Y%m%d')

#アップロード用ファイル名
upload_file = f'アップロード用ファイル/upload_{today}.txt'

read_excel_prog_per = 1000    #進捗表示用
GUI.var.set(0)                #進捗バーリセット


def FileExport_func():
    global today
    global upload_file

    # メール内容のアウトプットファイルを読み込む
    output_df = pd.read_excel(f'mail_import/import_{today}.xlsx')

    # SKUリストのシート名一覧を取得
    sku_list_file = pd.ExcelFile('SKUリスト.xlsx')
    sheet_names = sku_list_file.sheet_names

    #欠品表記追記用
    sku_list_file_for_write = load_workbook('SKUリスト.xlsx')

    read_excel_ind_prog = read_excel_prog_per / len(sheet_names)

    sku_list_df = {}
    for sheet_name in sheet_names:
        sku_list_df[sheet_name] = pd.read_excel(sku_list_file, sheet_name=sheet_name)
        countUp(read_excel_ind_prog)

    # 'import.xlsx'から処理対象のパラメータを取得
    target_products  = output_df['商品名'].tolist()
    target_powers    = output_df['度数'].tolist()
    target_tc_powers = output_df['乱視度数'] .tolist()
    target_tc_axis   = output_df['乱視軸'].tolist()
    target_bc        = output_df['BC'].tolist()
    target_add       = output_df['加入度数'].tolist()
    target_dia       = output_df['DIA'].tolist()
    target_color     = output_df['カラー'].tolist()
    target_lostflg   = output_df['欠品'].tolist()

    #欠品・解除SKUを格納する配列
    lost_sku_list    = []
    restock_sku_list = []

    #各パラメータを参照し、全てが合致するインデックス番号を取得
    for row_num, (product_name, power, tc_power, tc_axis, bc, add, dia, color, lostflg)\
    in enumerate(zip(target_products, target_powers ,target_tc_powers ,target_tc_axis, target_bc, target_add ,target_dia, target_color ,target_lostflg)):
        
        #パラメータ
        target_product_sheet_skus      = [] #SKU
        target_product_sheet_powers    = [] #度数
        target_product_sheet_tc_powers = [] #乱視度数
        target_product_sheet_tc_axis   = [] #乱視軸
        target_product_sheet_bc        = [] #BC
        target_product_sheet_add       = [] #ADD
        target_product_sheet_dia       = [] #DIA
        target_product_sheet_color     = [] #カラー

        #各パラメータごとの該当インデックス番号
        row_index_power                = []
        row_index_tc_power             = []
        row_index_tc_axis              = []
        row_index_bc                   = []
        row_index_add                  = []
        row_index_dia                  = []
        row_index_color                = []

        #全パラメータ統合
        intersection_list              = []
        
        #処理
        try:    #各パラメータを取得
            product_name_without_space = product_name.replace(' ','').replace('　','').replace('入り','')

            if product_name_without_space == 'ワンデーアキュビューオアシス乱視用':product_name_without_space = 'ワンデーアキュビューオアシス乱視用30枚'
            if product_name_without_space == 'ワンデーアキュビューモイスト乱視用':product_name_without_space = 'ワンデーアキュビューモイスト乱視用30枚'
            if product_name_without_space == 'エアオプティクスプラスハイドラグライド乱視用':product_name_without_space = 'エアオプティクスプラスハイドラグライド乱視用6枚'

            target_product_sheet = sku_list_df[product_name_without_space]

            target_product_sheet_for_write = sku_list_file_for_write[product_name_without_space]
            header_mapping = {cell.value: cell.column for cell in target_product_sheet_for_write[1]}
            lost_item_col                  = header_mapping['欠品']

            #print(target_product_sheet)
            print(product_name_without_space)

            try: 
                target_product_sheet_skus = target_product_sheet['SKU']
                #print(target_product_sheet_skus)
            except Exception as e:
                #print('SKU列がありません')
                continue

            try:
                target_product_sheet_powers = target_product_sheet['度数']
                row_index_power             = target_product_sheet_powers[target_product_sheet_powers == float(power)].index.tolist()
                test_mes = target_product_sheet_powers[target_product_sheet_powers == float(power)]
                print(test_mes)
                print(f'row_index_power={row_index_power}')
                #print(target_product_sheet_powers)
            except ValueError:
                target_product_sheet_powers = None
                print('度数の形式が正しくありません')
            except Exception as e:
                target_product_sheet_powers = None
                #print('度数列がありません')

            try:
                target_product_sheet_tc_powers = target_product_sheet['乱視度数']
                row_index_tc_power             = target_product_sheet_tc_powers[target_product_sheet_tc_powers == tc_power].index.tolist()
                test_mes2 = target_product_sheet_tc_powers[target_product_sheet_powers == float(tc_power)]
                print(test_mes2)
                print(f'row_index_tc_power={row_index_tc_power}')
                print(target_product_sheet_tc_powers)
            except Exception as e:
                target_product_sheet_tc_powers = None
                #print('乱視度数列がありません')

            try:
                target_product_sheet_tc_axis = target_product_sheet['乱視軸']
                row_index_tc_axis            = target_product_sheet_tc_axis[target_product_sheet_tc_axis == int(tc_axis)].index.tolist()
                print(f'row_index_tc_axis ={row_index_tc_axis}')
                print(target_product_sheet_tc_axis)
            except Exception as e:
                target_product_sheet_tc_axis = None
                #print('乱視軸列がありません')

            try:
                target_product_sheet_bc = target_product_sheet['BC']
                row_index_bc            = target_product_sheet_bc[target_product_sheet_bc == float(bc)].index.tolist()
                #print(target_product_sheet_bc)
            except Exception as e:
                target_product_sheet_bc = None
                #print('BC列がありません')

            try:
                target_product_sheet_add = target_product_sheet['加入度数']
                row_index_add            = target_product_sheet_add[target_product_sheet_add == float(add)].index.tolist()
                #print(target_product_sheet_add)
            except Exception as e:
                target_product_sheet_add = None
                #print('加入度数列がありません')

            try:
                target_product_sheet_dia = target_product_sheet['DIA']
                row_index_dia            = target_product_sheet_dia[target_product_sheet_dia == float(dia)].index.tolist()
                #print(target_product_sheet_dia)
            except Exception as e:
                target_product_sheet_dia = None
                #print('DIA列がありません')

            try:
                target_product_sheet_color = target_product_sheet['カラー']
                row_index_color            = target_product_sheet_dia[target_product_sheet_color == color].index.tolist()
                #print(target_product_sheet_color)
            except Exception as e:
                target_product_sheet_color = None
                #print('カラー列がありません')

            if row_index_power != []:
                marged = [value for value in row_index_power]
                print(marged)

                if row_index_tc_power != []:
                    marged = [value for value in row_index_power if value in row_index_tc_power]
                if row_index_tc_axis != []:
                    marged = [value for value in marged if value in row_index_tc_axis]
                if row_index_bc != []:
                    marged = [value for value in marged if value in row_index_bc]
                if row_index_add != []:
                    marged = [value for value in marged if value in row_index_add]
                if row_index_dia != []:
                    marged = [value for value in marged if value in row_index_dia]
                if row_index_color != []:
                    marged = [value for value in marged if value in row_index_color]
                
                intersection_list  = marged

            for row_loop_num in intersection_list:
                if lostflg == '欠品':
                    lost_sku_list.append(target_product_sheet_skus.iloc[row_loop_num])

                    #SKUリストに'欠品'を追加する
                    target_product_sheet_for_write.cell(row=row_loop_num + 2, column=lost_item_col, value= '欠品')
                else:
                    restock_sku_list.append(target_product_sheet_skus.iloc[row_loop_num])

                    #SKUリストの'欠品'を削除する
                    target_product_sheet_for_write.cell(row=row_loop_num + 2, column=lost_item_col, value= '')

        except Exception as e:
            #print(f'{product_name}：該当するシートがありません')
            continue
    
    #SKUリストの欠品表記を保存
    sku_list_file_for_write.save('SKUリスト.xlsx')

    # 欠品アップロード用ファイルのテンプレートを取得
    wb = Workbook()
    ws = wb.active

    template = pd.read_excel('Template.xlsx', header = None)

    for row in template.values: #ヘッダーを記述
        ws.append([row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11]])

    lost_item_rows = -1  #欠品商品の行数格納用

    for row, value in enumerate(lost_sku_list, start=4):
        ws.cell(row, 1, value)  #SKUを転記
        ws.cell(row, 3, 0)      #在庫数列に'0'を記入
        lost_item_rows = row - 4

    for row, value in enumerate(restock_sku_list, start=lost_item_rows + 5):    #欠品商品の行の直後から記述
        ws.cell(row, 1, value)  #SKUを転記
        ws.cell(row, 3, 10000)  #在庫数列に'10000'を記入
    
    with open(upload_file, 'w') as f :
        for write_row in ws.iter_rows(values_only = True):
            line = '\t'.join(str(cell) if not pd.isna(cell) else '' for cell in write_row) + '\n'
            f.write(line)

    #wb.save(f'アップロード用ファイル/upload_{today}.xlsx')

    #print('欠品=',lost_sku_list)
    #print('解除=',restock_sku_list)

    GUI.finished_progress = True
    countUp(1000-GUI.var.get())